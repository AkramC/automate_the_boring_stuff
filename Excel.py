import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb=openpyxl.load_workbook('D:\\files\\example1.xlsx')
print(wb.sheetnames)
sheet=wb.get_sheet_by_name('Sheet1')
print(sheet)
print(sheet['A1'].value)
print(sheet['B1'].value)
print(sheet.cell(row=1,column=1).value)
maxrow=sheet.max_row
maxcolumn=sheet.max_column
print(maxrow)
print(maxcolumn)
for i in range(1,8,1):
    for j in range(1,4,1):
        print(sheet.cell(row=i,column=j).value)

print(column_index_from_string('C'))
print(get_column_letter(maxcolumn))